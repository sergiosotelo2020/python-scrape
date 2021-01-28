import csv
import chilkat
import sys
import re
import os
import json
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait  # for implicit and explict waits
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from datetime import date
from selenium.webdriver.support.ui import WebDriverWait
import selenium.webdriver
from PIL import Image
import pytesseract
import time
from datetime import datetime
from openpyxl import load_workbook
import xlrd
import pandas as pd
from xlsxwriter.workbook import Workbook

option = webdriver.ChromeOptions()
option.add_argument('headless')
today = date.today()
# now = datetime.now()
current_time = today.strftime("%m%d%y")
print(current_time)
file_name = input("Enter the excel file name: ")
# driver = webdriver.Chrome('chromedriver', options = option)
driver = selenium.webdriver.Chrome()
ComapnyName = "Appex Corporate Solutions"
GemUsername = "Appex_corp22"
GemPassword = "Mybox@005"

base_url = 'https://sso.gem.gov.in/ARXSSO/oauth/login'
pp = 0
while pp < 10:
    driver.get(base_url)
    time.sleep(5)
    driver.implicitly_wait(5)
    element = driver.find_element_by_id("captcha1")
    location = element.location
    size = element.size
    driver.save_screenshot("pageImage.png")

    # crop image
    x = location['x']
    y = location['y']
    width = location['x']+size['width']
    height = location['y']+size['height']
    im = Image.open('pageImage.png')
    im = im.crop((int(x), int(y), int(width), int(height)))
    im.save('element.png')
    pytesseract.pytesseract.tesseract_cmd = r'C:/Program Files/Tesseract-OCR/tesseract.exe'
    #pytesseract.pytesseract.tesseract_cmd = r'C:/Users/Master/AppData/Local/tesseract.exe'
    text = pytesseract.image_to_string(Image.open("element.png"))
    print(text)
    # element=driver.find_element_by_xpath('//span[@class="last-page"]')

    print("----------start login---------------")
    user_id = driver.find_element_by_id('loginid')
    user_id.send_keys(GemUsername)
    captcha_text = driver.find_element_by_id('captcha_math')
    captcha_text.send_keys(text)
    print("--------Please insert key manually--------------")
    time.sleep(1)
    submit_button=driver.find_elements_by_css_selector("button.btn-nov")[0]
    submit_button.click()

    time.sleep(3)
    try:
        driver.find_element_by_id('password') 
        print('captcha bypass successfully')
        break
    except:
        pp += 1
        print('Sorry captcha')

password = driver.find_element_by_id('password')
password.send_keys(GemPassword)
time.sleep(1)
submit_button2 = driver.find_element_by_xpath('//button[@type="submit"]')
submit_button2.click()
print('login successfully')
driver.maximize_window()

#Read excel file
time.sleep(8)

edit_urls = []
statuss = []
errors = []

print('--------------read excel file-------------------')
workbook = load_workbook(filename=file_name)
# sheet_name = workbook.sheetnames
# print(sheet_name)
sheet = workbook.active
urls = []
x = 0
for i in sheet['G']: 
    url = i.value  
    if (url == None):
        break
    urls.append(url)
urls.remove("Direct Url")
x = len(urls)
print(x)
# try:
#     close_button = driver.find_element_by_id("flox-chat-close")
#     close_button.click()
# except NoSuchElementException:
#     print("don't exist such element")
loc = file_name
 
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
print(urls)
# sheet1 = xlrd.open_workbook("Web2Pair.xlsx").sheet_by_index(0) 

key = 1
for url in urls:
    
    driver.get(url)
    print('-----------redirect to edit page---------------------')
    time.sleep(15)
    try:
        driver.find_element_by_id("flox-chat-close")
        close_button = driver.find_element_by_id("flox-chat-close")
        try:
            close_button.click()
            
        except:
            print('Chat box N/A')
    except:
        print('Chat box N/A')

    i = 0

    rows = sheet.row_values(key)
                
    ProductCatalogID = rows[0]
    # print(ProductCatalogID)
    Model = rows[1]
    Category = rows[2]
    Brand = rows[3]
    Price_limit = rows[4]
    url = rows[5]
    Direct_url = rows[6]
    Authorization_no = rows[7]
    # print(Authorization_no)
    Authorization__agency = rows[8]
    # print(Authorization__agency)
    Authorization_date = int(rows[9])
    # print(Authorization_date)
    From = int(rows[10])
    # print(From)
    To = int(rows[11])
    # print(To)
    Country_of_origin = rows[12]
    sku = str(rows[13])
    # print(sku)
    hsn = int(rows[14])
    # print(hsn)
    mrp = int(rows[15])
    Offer_price = str(rows[16])
    # print(Offer_price)
    try:
        Pincodes = int(rows[17])
    except:
        Pincodes = ''
    Disticts = rows[18]
    # print(Disticts)
    State = rows[19]
    Current_stock = int(rows[20])
    mqpc = int(rows[21])
    # print(mqpc)
    lead_time = int(rows[22])
    edit_url = rows[23]
    status = rows[24]
    # print(status)
    workbook.close() 
    if float(rows[4]) > float(rows[16]):
        errors.append('OFfer price is lower than limit')
        print('Offer price is lower than limit')
        key += 1
        continue

    


    
    while i < 3:
        try:
            driver.find_element_by_class_name("popup-footer")
            popup = driver.find_element_by_class_name("popup-footer")
            popup_button = popup.find_elements_by_css_selector("*")[0]
            popup_button.click()
            time.sleep(10)

            try:
                print('Start insert')                    
                
                
                time.sleep(2)
                # chunks = [sku[i:i+1] for i in range(0, len(sku), 1)]
                # print('chunks')
                # print(chunks)
                # if chunks == '-':
                


                try:
                    sku_value = driver.find_elements_by_xpath('//div[@class="row"]/div[@class="col-sm-6"]/input[@type="text"]')[2]
                    print("Exit SKU1")
                    # actions = ActionChains(driver)
                    # actions.move_to_element(sku_value).perform()
                    # sku_value.click()
                    sku_value.send_keys(sku)
                    try:
                        auth_no = driver.find_elements_by_xpath('//div[@class="row"]/div[@class="col-sm-6"]/input[@type="text"]')[0]
                        print('Exit Auth No')
                        auth_no.send_keys(Authorization_no)
                    except:
                        print('Auth No N/A')
                    try:
                        auth_agency = driver.find_elements_by_xpath('//div[@class="row"]/div[@class="col-sm-6"]/input[@type="text"]')[1]
                        print('Exit Auth agency')
                        auth_agency.send_keys(Authorization__agency)
                    except:
                        print('Auth agency N/A')
                    
                except:
                    sku_value = driver.find_element_by_xpath('//div[@class="row"]/div[@class="col-sm-6"]/input[@type="text"]')
                    print('Exit SKU2')
                    # actions = ActionChains(driver)
                    # actions.move_to_element(sku_value).perform()
                    # sku_value.click()
                    sku_value.send_keys(sku)
                # Authorization data
                xl_date = Authorization_date
                datetime_date = xlrd.xldate_as_datetime(xl_date, 0)
                date_object = datetime_date.date()

                year = date_object.year
                print(year)
                month = date_object.month
                print(month)
                kk = month - 1
                day = date_object.day
                print(day)
                try:
                    print("date handle start")
                    try:
                        date_buttons = driver.find_elements_by_xpath('//span/button[@class="btn btn-default"]')
                        date_buttons[0].click()
                    except:
                        print('no calendar button')
                    print("date 2")
                    try:
                        tt_button = driver.find_element_by_xpath('//th/button[@class="btn btn-default btn-sm uib-title"]')
                        tt_button.click()
                        date_tmp = driver.find_element_by_xpath('//button/strong[@class="ng-binding"]').text
                        date_tmp = int(date_tmp)
                        print(date_tmp)
                        if year == date_tmp:
                            print('OK')
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue
    

                        elif date_tmp > year:
                            nn = date_tmp - year
                            for xx in range(nn):
                                driver.find_element_by_xpath('//button[@class="btn btn-default btn-sm pull-left uib-left"]').click()
                                print(xx)
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue
         
                            
                        else:
                            nn = year - date_tmp
                            for xx in range(nn):
                                driver.find_element_by_xpath('//button[@class="btn btn-default btn-sm pull-right uib-right"]').click()
                                print(xx)
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue
                    except:
                        print('no button')
                        
                    
      
                except:
                    print('Authorize date Error')

                # From data
                xl_date = From
                datetime_date = xlrd.xldate_as_datetime(xl_date, 0)
                date_object = datetime_date.date()

                year = date_object.year
                print(year)
                month = date_object.month
                print(month)
                kk = month - 1
                day = date_object.day
                print(day)
                try:
                    print("date handle start")
                    try:
                        date_buttons = driver.find_elements_by_xpath('//span/button[@class="btn btn-default"]')
                        date_buttons[1].click()
                    except:
                        print('no calendar button')
                    print("date 2")
                    try:
                        driver.find_element_by_xpath('//th/button[@class="btn btn-default btn-sm uib-title"]').click()
                        date_tmp = driver.find_element_by_xpath('//button/strong[@class="ng-binding"]').text
                        date_tmp = int(date_tmp)
                        print(date_tmp)
                        if year == date_tmp:
                            print('OK')
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue


                        elif date_tmp > year:
                            nn = date_tmp - year
                            for xx in range(nn):
                                driver.find_element_by_xpath('//button[@class="btn btn-default btn-sm pull-left uib-left"]').click()
                                print(xx)
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue
  
                            
                        else:
                            nn = year - date_tmp
                            for xx in range(nn):
                                driver.find_element_by_xpath('//button[@class="btn btn-default btn-sm pull-right uib-right"]').click()
                                print(xx)
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue
                    except:
                        print('no button')
   
                except:
                    print('Aithorize date Error')
                
                # To data
                xl_date = To
                datetime_date = xlrd.xldate_as_datetime(xl_date, 0)
                date_object = datetime_date.date()

                year = date_object.year
                print(year)
                month = date_object.month
                print(month)
                kk = month - 1
                day = date_object.day
                print(day)
                try:
                    print("date handle start")
                    try:
                        date_buttons = driver.find_elements_by_xpath('//span/button[@class="btn btn-default"]')
                        date_buttons[2].click()
                    except:
                        print('no calendar button')
                    print("date 2")
                    try:
                        driver.find_element_by_xpath('//th/button[@class="btn btn-default btn-sm uib-title"]').click()
                        date_tmp = driver.find_element_by_xpath('//button/strong[@class="ng-binding"]').text
                        date_tmp = int(date_tmp)
                        print(date_tmp)
                        if year == date_tmp:
                            print('OK')
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue
                   

                        elif date_tmp > year:
                            nn = date_tmp - year
                            for xx in range(nn):
                                driver.find_element_by_xpath('//button[@class="btn btn-default btn-sm pull-left uib-left"]').click()
                                print(xx)
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue
                       
                            
                        else:
                            nn = year - date_tmp
                            for xx in range(nn):
                                driver.find_element_by_xpath('//button[@class="btn btn-default btn-sm pull-right uib-right"]').click()
                                print(xx)
                            month_button = driver.find_elements_by_xpath('//td[@class="uib-month text-center ng-scope"]/button')[kk]
                            month_button.click()
                            day_buttons = driver.find_elements_by_xpath('//td[@class="uib-day text-center ng-scope"]/button/span')
                            for day_button in day_buttons:
                                dd = day_button.text
                                dd = int(dd)
                                if dd == day:
                                    day_button.click()
                                    break
                                else:
                                    continue
                    except:
                        print('no button')
                        
                    
                except:
                    print('Aithorize date Error')
                
                # third_part = driver.find_element_by_class_name("stock-section-fieldset")
                try:
                    country = driver.find_elements_by_css_selector("input.input-xs")[0]
                    print("Exit country")
                    country.send_keys(Country_of_origin)
                    time.sleep(2)
                    driver.find_element_by_class_name("ui-select-choices-row-inner").click()
                except:
                    print('Country N/A')
                
                time.sleep(2)
                
                try:
                    hsn_value = driver.find_element_by_xpath('//div[@class="row"]/div[@class="col-sm-6 wsp-tool-tip-wrap"]/input[@type="text"]')
                    print("Exit HSN")
                    hsn_value.send_keys(hsn)
                    
                except:
                    print("HSN N/A")
                time.sleep(1)
                
                try:
                    mrp_value = driver.find_element_by_xpath('//div[@class="row ng-scope"]/div[@class="col-sm-6 tool-tip-wrap"]/input[@type="number"]')
                    print("Exit MRP")
                    mrp_value.send_keys(mrp)
                    
                except:
                    print('MRP N/A')
                time.sleep(1)

                try:
                    offer_price_value = driver.find_element_by_xpath('//div[@class="row"]/div[@class="col-sm-5 wsp-tool-tip-wrap"]/input[@type="number"]')
                    offer_price_value.send_keys(Offer_price)
                    print("Exit offer price")
                except:
                    print('Offer price N/A')
                
                time.sleep(2)
                try:
                    disticts_value = driver.find_elements_by_xpath('//div[@class="ui-select-container ui-select-multiple ui-select-bootstrap dropdown form-control ng-pristine ng-untouched ng-valid ng-scope ng-empty"]/div/input[@type="search"]')
                    print("Exit distict")
                    disticts_value[0].send_keys(Disticts)
                except:
                    print('Distict N/A')
                    
                time.sleep(2)
                if Disticts == '':
                    print("Disticts N/A")
                else:
                    try:
                        driver.find_elements_by_xpath('//div[@class="ui-select-choices-row ng-scope active"]/span[@class="ui-select-choices-row-inner"]')[1].click()
                    except:
                        print('No dropdown')
                time.sleep(2)
                print("1")
                try:
                    if Disticts == '':
                        pincode = driver.find_elements_by_xpath('//div[@class="panel-body"]/div[@class="ui-select-container ui-select-multiple ui-select-bootstrap dropdown form-control ng-pristine ng-untouched ng-valid ng-scope ng-empty"]/div/input[@type="search"]')[1]
                        print("Exit pincode")
                        pincode.send_keys(Pincodes)
                        if Pincodes == '':
                            print("Pincodes N/A")
                        else:
                            try:
                                driver.find_elements_by_xpath('//div[@class="ui-select-choices-row ng-scope active"]/span[@class="ui-select-choices-row-inner"]')[1].click()
                            except:
                                print('Pincode N/A')
                    else:
                        pincode = driver.find_element_by_xpath('//div[@class="panel-body"]/div[@class="ui-select-container ui-select-multiple ui-select-bootstrap dropdown form-control ng-pristine ng-untouched ng-valid ng-scope ng-empty"]/div/input[@type="search"]')
                        print("Exit Pincode 2")
                        pincode.send_keys(Pincodes)
                        if Pincodes == '':
                            print("Pincodes N/A")
                        else:
                            try:
                                driver.find_element_by_xpath('//div[@class="ui-select-choices-row ng-scope active"]/span[@class="ui-select-choices-row-inner"]').click()
                            except:
                                print('No dropdown')
                except:
                    pincode = driver.find_elements_by_xpath('//div[@class="panel-body"]/div[@class="ui-select-container ui-select-multiple ui-select-bootstrap dropdown form-control ng-pristine ng-untouched ng-valid ng-scope ng-empty"]/div/input[@type="search"]')[2]
                    print("Exit pincode 3")
                    pincode.send_keys(Pincodes)
                    if Pincodes == '':
                        print("Pincodes N/A")
                    else:
                        try:
                            driver.find_element_by_xpath('//div[@class="ui-select-choices-row ng-scope active"]/span[@class="ui-select-choices-row-inner"]').click()
                        except:
                            print('No dropdown')                  
                time.sleep(3)
                try:
                    stock = driver.find_elements_by_xpath('//div[@class="row"]/div[@class="col-sm-6 tool-tip-wrap"]/input[@type="number"]')
                    current_stock_input = stock[0]
                    current_stock_input.send_keys(Current_stock)
                    mqpc_input = stock[1]
                    mqpc_input.send_keys(mqpc)
                    lead_time_input = stock[2]
                    lead_time_input.send_keys(lead_time)
                except:
                    print('Stock Values N/A')
                if State == '':
                    print('State N/A')
                else:
                    states = driver.find_elements_by_xpath('//td/span/input[@type="checkbox"]')
                    try:
                        liElement = driver.find_elements_by_xpath('//div[@class="panel-default ng-scope ng-isolate-scope panel panel-open"]/div/h4/a/span/div/div[@class="input-group-item tab-heading"]')[0]
                        driver.execute_script("arguments[0].scrollIntoView(true);", liElement)
                    except:
                        print("Ok")
                    for state in states:
                        try:
                            state.click()
                        except:
                            print('State N/A')
                
                print("Save Step")
                time.sleep(5)
                try:
                    save_button = driver.find_element_by_xpath('//button[@class="button make-model-submit ng-scope ng-isolate-scope"]')
                    save_button.click()
                    print('Save button click')
                    time.sleep(10)
                    
                    review_button = driver.find_element_by_xpath('//a[@class="button success-button"]')
                    review_button.click()
                    print("Button click")
                    errors.append('')
                except:
                    print('here error')
                    try:
                        try:
                            ddd = driver.find_elements_by_xpath('//div[@class="input-group-item fa fa-2 circle-right fa-chevron-circle-right"]')[2]
                        except:
                            ddd = driver.find_elements_by_xpath('//div[@class="input-group-item fa fa-2 circle-right fa-chevron-circle-right"]')[1]
                        ddd.click()
                        time.sleep(2)
                        error = ''
                        errors_tmp = driver.find_elements_by_xpath('//span[@class="tool-tip ng-binding ng-scope"]')
                        for error_tmp in errors_tmp:
                            error += error_tmp.text + ","
                        errors.append(error)
                    except:
                        errors.append('Already exit')
                time.sleep(5)
                
                try:
                    driver.switch_to.window(driver.window_handles[1])
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    print("Switch Browser Success")
                except:
                    print('No save button')
                time.sleep(1)


                

                
                edit_url = driver.current_url
                print(edit_url)
                edit_urls.append(edit_url)
                time.sleep(2)
                try:
                    publish_button = driver.find_element_by_xpath('//input[@class="button success-button center-block ng-isolate-scope"]')
                    publish_button.click()

                    time.sleep(2)
                    publish_button2 = driver.find_element_by_xpath('//div/div[@class="btn btn-primary"]')
                    publish_button2.click()
                    statu = 'Published'
                    
                except:
                    print('Publish Button N/A')
                    try:   
                        agree = driver.find_element_by_xpath('//div[@class="text-center agree ng-binding ng-scope"]/input[@type="checkbox"]')   
                        agree.click()
                    except:
                        print("No check box")
                    publish_button = driver.find_element_by_xpath('//input[@class="button success-button center-block ng-isolate-scope"]')
                    publish_button.click()
                    time.sleep(3)
                    try:
                        publish_button2 = driver.find_element_by_xpath('//div/div[@class="btn btn-primary"]')
                        publish_button2.click()
                        time.sleep(5)
                        # statu = driver.find_elements_by_xpath('//tr/td[@id="status"]')[0].text
                        statu = 'Published'
                        print(statu)
                    except:
                        err_len = len(errors)
                        statu = ''
                        kkk = key -1
                        if errors[kkk] == '':
                            statu = 'Already exits'
                        else:
                            statu = 'Errors Occur'
                        
                print(statu)
                statuss.append(statu)

                time.sleep(3)
                print("Error Status")
                
                print(errors)
                break

            except:
                
                print('Big Error Occur')

        except:
            print('No popup')
            time.sleep(5)
            driver.refresh()
            
            i += 1
            continue

    time.sleep(5)
    print("Done" + str(key))
    print('-----------------------------------------------------')
    key += 1
    continue  

wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
# sheet1 = xlrd.open_workbook("Web2Pair.xlsx").sheet_by_index(0) 
tt = open("T.csv", 'w', newline="")
col = csv.writer(tt) 

for row in range(sheet.nrows): 
    col.writerow(sheet.row_values(row))  
# df = pd.DataFrame(pd.read_csv("T.csv")) 
# df
tt.close()
csvv = chilkat.CkCsv()
csvv.put_HasColumnNames(True)
time.sleep(1)

success = csvv.LoadFile("T.csv")
if (success != True):
    print(csvv.lastErrorText())
    sys.exit()
k = 0
ee = len(edit_urls)
print('urls:' + str(ee))
ss = len(statuss)
print('status:' + str(ss))
er = len(errors)
print('error:' + str(er))

if ee < ss:
    ee = ss

while k < ee:
    success = csvv.SetCell(k,23,edit_urls[k])
    success = csvv.SaveFile("V.csv")
    success = csvv.SetCell(k,24,statuss[k])
    success = csvv.SaveFile("V.csv")
    success = csvv.SetCell(k,25,errors[k])
    success = csvv.SaveFile("V.csv")
    k += 1


if (success != True):
    print(csvv.lastErrorText())
time.sleep(1)
csvfile = "V.csv"

workbook = Workbook("result" + file_name)
worksheet = workbook.add_worksheet()
with open(csvfile, 'rt', encoding='utf8') as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader):
        for c, col in enumerate(row):
            worksheet.write(r, c, col)
workbook.close()  

time.sleep(2)
os.remove("T.csv")
os.remove("V.csv")

print("All done!!!")
print("------------------------------------")
print("Please check result" + file_name)